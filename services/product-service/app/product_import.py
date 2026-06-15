import json
import os
import re
from datetime import datetime
from io import BytesIO

import requests
from fastapi import HTTPException, UploadFile
from sqlalchemy.orm import Session

from product_models import Category, Product, ProductImportJob
from product_serializers import (
    _product_archived_at,
    _product_description,
    _product_id,
    _product_name,
    _product_sku,
    _set_product_archived_at,
)
from shared.config import settings
from shared.money import as_money, money_json
from shared.redis_client import redis_client

PRODUCT_CACHE_KEY = "products:all"
CATEGORY_ALIASES = {
    "accessories": "Accesorii",
    "accesories": "Accesorii",
    "accesorii": "Accesorii",
    "keyboards": "Periferice",
    "mice": "Periferice",
    "periferice": "Periferice",
}
INVENTORY_SERVICE_URL = os.getenv("INVENTORY_SERVICE_URL", "http://inventory-service:8000")
IMPORT_REQUIRED_HEADERS = {"sku", "name", "description", "price", "category", "stock", "active"}
IMPORT_HEADER_ALIASES = {
    "cod": "sku",
    "sku": "sku",
    "nume": "name",
    "name": "name",
    "descriere": "description",
    "description": "description",
    "pret": "price",
    "price": "price",
    "categorie": "category",
    "category": "category",
    "stoc": "stock",
    "stock": "stock",
    "activ": "active",
    "active": "active",
    "operatie": "operation",
    "operation": "operation",
}


def normalize_sku(value: str) -> str:
    normalized = re.sub(r"[^A-Z0-9]+", "-", value.strip().upper()).strip("-")
    if not normalized:
        raise HTTPException(status_code=400, detail="SKU must contain letters or digits")
    return normalized[:80]


def generate_sku(name: str, product_id: int) -> str:
    base = normalize_sku(name)[:64]
    return f"{base}-{product_id}"


def canonical_category_name(value: str) -> str:
    name = " ".join(value.split())
    return CATEGORY_ALIASES.get(name.casefold(), name)


def _parse_import_decimal(value, *, field_name: str):
    text = str(value).strip().replace(",", ".")
    if not text:
        raise HTTPException(status_code=400, detail=f"{field_name} is required")
    try:
        parsed = as_money(text)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"{field_name} must be a valid amount") from exc
    if parsed <= 0:
        raise HTTPException(status_code=400, detail=f"{field_name} must be greater than 0")
    return parsed


def _parse_import_stock(value) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        stock = int(float(str(value).strip().replace(",", ".")))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="stock must be an integer") from exc
    if stock < 0:
        raise HTTPException(status_code=400, detail="stock must be >= 0")
    return stock


def _parse_import_active(value) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return True
    if text in {"1", "true", "yes", "da", "y"}:
        return True
    if text in {"0", "false", "no", "nu", "n"}:
        return False
    raise HTTPException(status_code=400, detail="active must be true/false")


def _parse_import_operation(value) -> str | None:
    text = str(value or "").strip().lower()
    if not text:
        return None
    if text not in {"upsert", "archive"}:
        raise HTTPException(status_code=400, detail="operation must be upsert or archive")
    return text


def _sync_inventory_bulk(items: list[dict]) -> None:
    if not items:
        return
    response = requests.post(
        f"{INVENTORY_SERVICE_URL}/inventory/internal/bulk-seed",
        json={"items": items},
        timeout=10,
        headers={"X-Internal-Api-Token": settings.internal_api_token or ""},
    )
    if response.status_code != 200:
        raise HTTPException(status_code=502, detail="Inventory service rejected bulk stock update")


def _enforce_import_upload(file: UploadFile, file_bytes: bytes) -> None:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Upload an .xlsx file")
    if len(file_bytes) > settings.admin_import_max_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"Excel file exceeds {settings.admin_import_max_bytes} bytes",
        )
    if file.content_type not in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
    }:
        raise HTTPException(status_code=400, detail="Unsupported upload content type")


def _load_import_rows(file_bytes: bytes) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    raw_headers = [str(cell or "").strip() for cell in rows[0]]
    normalized_headers: list[str] = []
    for header in raw_headers:
        key = IMPORT_HEADER_ALIASES.get(header.casefold())
        normalized_headers.append(key or header.casefold())

    missing = sorted(IMPORT_REQUIRED_HEADERS - set(normalized_headers))
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(missing)}",
        )

    payload_rows: list[dict] = []
    for index, raw_row in enumerate(rows[1:], start=2):
        values = dict(zip(normalized_headers, raw_row))
        if all(value in (None, "") for value in values.values()):
            continue
        payload_rows.append({"row_number": index, **values})

    if not payload_rows:
        raise HTTPException(status_code=400, detail="Excel file does not contain product rows")
    return payload_rows


def _build_import_preview(rows: list[dict], db: Session) -> dict:
    categories_by_name = {category.name.casefold(): category for category in db.query(Category).all()}
    existing_products = {_product_sku(product): product for product in db.query(Product).all()}
    preview_rows: list[dict] = []
    summary = {"create": 0, "update": 0, "archive": 0, "skip": 0, "error": 0}

    for row in rows:
        result = {
            "row_number": row["row_number"],
            "sku": str(row.get("sku") or "").strip(),
            "name": str(row.get("name") or "").strip(),
            "category": str(row.get("category") or "").strip(),
            "stock": row.get("stock"),
        }
        try:
            sku = normalize_sku(result["sku"])
            name = str(row.get("name") or "").strip()
            description = str(row.get("description") or "").strip()
            category_name = canonical_category_name(str(row.get("category") or "").strip())
            operation = _parse_import_operation(row.get("operation"))
            active = _parse_import_active(row.get("active"))
            price = _parse_import_decimal(row.get("price"), field_name="price")
            stock = _parse_import_stock(row.get("stock"))

            if not name:
                raise HTTPException(status_code=400, detail="name is required")
            if not description:
                raise HTTPException(status_code=400, detail="description is required")
            if not category_name:
                raise HTTPException(status_code=400, detail="category is required")

            product = existing_products.get(sku)
            requested_action = operation or ("archive" if not active else "upsert")
            category_exists = category_name.casefold() in categories_by_name

            if requested_action == "archive":
                if not product:
                    raise HTTPException(status_code=400, detail="cannot archive unknown SKU")
                action = "skip" if _product_archived_at(product) else "archive"
                reason = "already archived" if _product_archived_at(product) else None
            elif product:
                action = "update"
                reason = "will restore archived product" if _product_archived_at(product) else None
            else:
                action = "create"
                reason = None

            result.update(
                {
                    "sku": sku,
                    "name": name,
                    "description": description,
                    "price": money_json(price),
                    "category": category_name,
                    "stock": 0 if stock is None and action == "create" else stock,
                    "active": active,
                    "operation": requested_action,
                    "action": action,
                    "category_action": "existing" if category_exists else "create",
                    "message": reason,
                }
            )
            summary[action] += 1
        except HTTPException as exc:
            result["action"] = "error"
            result["message"] = exc.detail
            summary["error"] += 1
        preview_rows.append(result)

    return {"summary": summary, "rows": preview_rows}


def _apply_import(preview: dict, db: Session, *, filename: str, created_by: str) -> dict:
    if preview["summary"]["error"]:
        raise HTTPException(status_code=400, detail="Fix import errors before apply")

    categories_by_name = {category.name.casefold(): category for category in db.query(Category).all()}
    inventory_updates: list[dict] = []
    applied = {"created": 0, "updated": 0, "archived": 0, "restored": 0, "stock_updates": 0}

    for row in preview["rows"]:
        if row["action"] == "skip":
            continue

        category = categories_by_name.get(row["category"].casefold())
        if not category:
            category = Category(
                name=row["category"],
                description=f"Importată din Excel pe {datetime.utcnow().date().isoformat()}",
            )
            db.add(category)
            db.flush()
            categories_by_name[category.name.casefold()] = category

        product = db.query(Product).filter(Product.sku == row["sku"]).first()
        if row["action"] == "archive":
            assert product is not None
            _set_product_archived_at(product, datetime.utcnow())
            db.add(product)
            inventory_updates.append({"product_id": _product_id(product), "stock": 0})
            applied["archived"] += 1
            applied["stock_updates"] += 1
            continue

        if product:
            restored = _product_archived_at(product) is not None
            product.name = row["name"]
            product.description = row["description"]
            product.price = as_money(row["price"])
            product.category_id = category.id
            _set_product_archived_at(product, None)
            db.add(product)
            applied["updated"] += 1
            if restored:
                applied["restored"] += 1
        else:
            product = Product(
                sku=row["sku"],
                name=row["name"],
                description=row["description"],
                price=as_money(row["price"]),
                category_id=category.id,
                archived_at=None,
            )
            db.add(product)
            db.flush()
            applied["created"] += 1

        if row["stock"] is not None:
            inventory_updates.append({"product_id": _product_id(product), "stock": int(row["stock"])})
            applied["stock_updates"] += 1

    db.commit()
    if inventory_updates:
        _sync_inventory_bulk(inventory_updates)
    redis_client.delete(PRODUCT_CACHE_KEY)
    result = {
        "message": "Import completed",
        "summary": applied,
        "rows": preview["rows"],
    }
    db.add(
        ProductImportJob(
            filename=filename[:255],
            summary_json=json.dumps(result["summary"]),
            created_by=created_by[:255],
        )
    )
    db.commit()
    return result


def build_import_template() -> BytesIO:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Produse"
    sheet.append(["sku", "name", "description", "price", "category", "stock", "active", "operation"])
    sheet.append(["DOCK-USBC-001", "USB-C Dock", "Dock cu HDMI si ethernet", 149.99, "Accesorii", 25, "true", "upsert"])
    sheet.append(["OLD-MOUSE-001", "Mouse vechi", "Produs retras", 49.00, "Periferice", 0, "false", "archive"])
    payload = BytesIO()
    workbook.save(payload)
    payload.seek(0)
    return payload


def build_products_export(products: list[Product], categories_by_id: dict[int, Category]) -> BytesIO:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Produse"
    sheet.append(["sku", "name", "description", "price", "category", "active", "archived_at"])
    for product in products:
        category = categories_by_id.get(product.category_id)
        archived_at = _product_archived_at(product)
        sheet.append(
            [
                _product_sku(product),
                _product_name(product),
                _product_description(product),
                money_json(product.price),
                category.name if category else "",
                "true" if archived_at is None else "false",
                archived_at.isoformat() if archived_at else "",
            ]
        )

    payload = BytesIO()
    workbook.save(payload)
    payload.seek(0)
    return payload
