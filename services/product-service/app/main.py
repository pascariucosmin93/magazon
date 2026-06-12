import os
import json
import re
from typing import Any, cast
from io import BytesIO
from datetime import datetime
from decimal import Decimal
from uuid import uuid4

from alembic.config import Config as AlembicConfig
from alembic import command as alembic_command
from fastapi import Depends, File, HTTPException, Query, Request, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy import Column, DateTime, ForeignKey, Integer, Numeric, String, Text, func
from sqlalchemy.orm import Session
from starlette.responses import StreamingResponse
import requests

from shared.auth import current_user_claims
from shared.config import settings
from shared.db import Base, SessionLocal, get_db
from shared.money import as_money, money_json
from shared.rate_limit import enforce_rate_limit
from shared.redis_client import redis_client
from shared.service_app import create_base_app


PRODUCT_CACHE_KEY = "products:all"
CATEGORY_ALIASES = {
    "accessories": "Accesorii",
    "accesories": "Accesorii",
    "accesorii": "Accesorii",
    "keyboards": "Periferice",
    "mice": "Periferice",
    "periferice": "Periferice",
}
INVENTORY_SERVICE_URL = os.getenv("INVENTORY_SERVICE_URL", "http://inventory-service:8000")
IMPORT_REQUIRED_HEADERS = {"sku", "name", "description", "price", "category", "stock", "active"}
IMPORT_HEADER_ALIASES = {
    "cod": "sku",
    "sku": "sku",
    "nume": "name",
    "name": "name",
    "descriere": "description",
    "description": "description",
    "pret": "price",
    "price": "price",
    "categorie": "category",
    "category": "category",
    "stoc": "stock",
    "stock": "stock",
    "activ": "active",
    "active": "active",
    "operatie": "operation",
    "operation": "operation",
}


class Category(Base):
    __tablename__ = "categories"

    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(255), unique=True, nullable=False, index=True)
    description = Column(Text, nullable=False, default="")
    created_at = Column(DateTime, default=datetime.utcnow)


class Product(Base):
    __tablename__ = "products"

    id = Column(Integer, primary_key=True, index=True)
    sku = Column(String(80), unique=True, nullable=False, index=True)
    name = Column(String(255), nullable=False)
    description = Column(Text, nullable=False)
    price = Column(Numeric(12, 2), nullable=False)
    category_id = Column(Integer, ForeignKey("categories.id"), nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    archived_at = Column(DateTime, nullable=True)


class ProductRequest(BaseModel):
    sku: str | None = Field(default=None, max_length=80)
    name: str = Field(min_length=1, max_length=255)
    description: str = Field(min_length=1)
    price: Decimal = Field(gt=0, decimal_places=2, max_digits=12)
    category_id: int | None = None


class CategoryRequest(BaseModel):
    name: str
    description: str


class ProductImportJob(Base):
    __tablename__ = "product_import_jobs"

    id = Column(Integer, primary_key=True, index=True)
    filename = Column(String(255), nullable=False)
    summary_json = Column(Text, nullable=False)
    created_by = Column(String(255), nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False, index=True)


def _run_migrations() -> None:
    service_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    cfg = AlembicConfig()
    cfg.set_main_option("script_location", os.path.join(service_dir, "migrations"))
    cfg.set_main_option("sqlalchemy.url", settings.database_url)
    alembic_command.upgrade(cfg, "head")


async def startup():
    _run_migrations()
    with SessionLocal() as db:
        if db.query(Category).count() == 0:
            db.add_all(
                [
                    Category(
                        name="Periferice",
                        description="Tastaturi, mouse-uri, căști și camere web",
                    ),
                    Category(
                        name="Accesorii",
                        description="Dock-uri, cabluri, suporturi și accesorii pentru birou",
                    ),
                ]
            )
            db.commit()

        categories = {category.name: category.id for category in db.query(Category).all()}
        if db.query(Product).count() == 0:
            db.add_all(
                [
                    Product(
                        sku="KEYBOARD-MECH-001",
                        name="Mechanical Keyboard",
                        description="Tactile RGB keyboard",
                        price=119.0,
                        category_id=categories.get("Periferice"),
                    ),
                    Product(
                        sku="MOUSE-GAMING-001",
                        name="Gaming Mouse",
                        description="Lightweight wireless mouse",
                        price=79.0,
                        category_id=categories.get("Periferice"),
                    ),
                    Product(
                        sku="DOCK-USBC-001",
                        name="USB-C Dock",
                        description="Dock with HDMI and ethernet",
                        price=149.0,
                        category_id=categories.get("Accesorii"),
                    ),
                ]
            )
            db.commit()
    redis_client.delete(PRODUCT_CACHE_KEY)


app = create_base_app("product-service", startup_hook=startup, check_db=True, check_redis=True)


def require_admin(claims: dict = Depends(current_user_claims)):
    if claims.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")
    return claims


def _product_id(product: Product) -> int:
    return cast(int, product.id)


def _product_sku(product: Product) -> str:
    return cast(str, product.sku)


def _product_name(product: Product) -> str:
    return cast(str, product.name)


def _product_description(product: Product) -> str:
    return cast(str, product.description)


def _product_category_id(product: Product) -> int | None:
    return cast(int | None, product.category_id)


def _product_archived_at(product: Product) -> datetime | None:
    return cast(datetime | None, product.archived_at)


def _set_product_sku(product: Product, value: str) -> None:
    cast(Any, product).sku = value


def _set_product_archived_at(product: Product, value: datetime | None) -> None:
    cast(Any, product).archived_at = value


def serialize_product(product: Product, categories_by_id: dict[int, Category]) -> dict:
    category = categories_by_id.get(_product_category_id(product))
    archived_at = _product_archived_at(product)
    return {
        "id": _product_id(product),
        "sku": _product_sku(product),
        "name": _product_name(product),
        "description": _product_description(product),
        "price": money_json(product.price),
        "category_id": _product_category_id(product),
        "category_name": category.name if category else None,
        "archived": archived_at is not None,
        "archived_at": archived_at.isoformat() if archived_at else None,
    }


def serialize_category(category: Category) -> dict:
    return {"id": category.id, "name": category.name, "description": category.description}


def serialize_import_job(job: ProductImportJob) -> dict:
    return {
        "id": job.id,
        "filename": job.filename,
        "summary": json.loads(cast(str, job.summary_json)),
        "created_by": job.created_by,
        "created_at": job.created_at.isoformat() if job.created_at else None,
    }


def normalize_sku(value: str) -> str:
    normalized = re.sub(r"[^A-Z0-9]+", "-", value.strip().upper()).strip("-")
    if not normalized:
        raise HTTPException(status_code=400, detail="SKU must contain letters or digits")
    return normalized[:80]


def generate_sku(name: str, product_id: int) -> str:
    base = normalize_sku(name)[:64]
    return f"{base}-{product_id}"


def canonical_category_name(value: str) -> str:
    name = " ".join(value.split())
    return CATEGORY_ALIASES.get(name.casefold(), name)


def _parse_import_decimal(value, *, field_name: str) -> Decimal:
    text = str(value).strip().replace(",", ".")
    if not text:
        raise HTTPException(status_code=400, detail=f"{field_name} is required")
    try:
        parsed = as_money(text)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"{field_name} must be a valid amount") from exc
    if parsed <= 0:
        raise HTTPException(status_code=400, detail=f"{field_name} must be greater than 0")
    return parsed


def _parse_import_stock(value) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        stock = int(float(str(value).strip().replace(",", ".")))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="stock must be an integer") from exc
    if stock < 0:
        raise HTTPException(status_code=400, detail="stock must be >= 0")
    return stock


def _parse_import_active(value) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return True
    if text in {"1", "true", "yes", "da", "y"}:
        return True
    if text in {"0", "false", "no", "nu", "n"}:
        return False
    raise HTTPException(status_code=400, detail="active must be true/false")


def _parse_import_operation(value) -> str | None:
    text = str(value or "").strip().lower()
    if not text:
        return None
    if text not in {"upsert", "archive"}:
        raise HTTPException(status_code=400, detail="operation must be upsert or archive")
    return text


def _sync_inventory_bulk(items: list[dict]) -> None:
    if not items:
        return
    response = requests.post(
        f"{INVENTORY_SERVICE_URL}/inventory/internal/bulk-seed",
        json={"items": items},
        timeout=10,
        headers={"X-Internal-Api-Token": settings.internal_api_token or ""},
    )
    if response.status_code != 200:
        raise HTTPException(status_code=502, detail="Inventory service rejected bulk stock update")


def _enforce_import_upload(file: UploadFile, file_bytes: bytes) -> None:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Upload an .xlsx file")
    if len(file_bytes) > settings.admin_import_max_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"Excel file exceeds {settings.admin_import_max_bytes} bytes",
        )
    if file.content_type not in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
    }:
        raise HTTPException(status_code=400, detail="Unsupported upload content type")


def _load_import_rows(file_bytes: bytes) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    raw_headers = [str(cell or "").strip() for cell in rows[0]]
    normalized_headers: list[str] = []
    for header in raw_headers:
        key = IMPORT_HEADER_ALIASES.get(header.casefold())
        normalized_headers.append(key or header.casefold())

    missing = sorted(IMPORT_REQUIRED_HEADERS - set(normalized_headers))
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(missing)}",
        )

    payload_rows: list[dict] = []
    for index, raw_row in enumerate(rows[1:], start=2):
        values = dict(zip(normalized_headers, raw_row))
        if all(value in (None, "") for value in values.values()):
            continue
        payload_rows.append({"row_number": index, **values})

    if not payload_rows:
        raise HTTPException(status_code=400, detail="Excel file does not contain product rows")
    return payload_rows


def _build_import_preview(rows: list[dict], db: Session) -> dict:
    categories_by_name = {
        category.name.casefold(): category
        for category in db.query(Category).all()
    }
    existing_products = {
        _product_sku(product): product
        for product in db.query(Product).all()
    }
    preview_rows: list[dict] = []
    summary = {"create": 0, "update": 0, "archive": 0, "skip": 0, "error": 0}

    for row in rows:
        result = {
            "row_number": row["row_number"],
            "sku": str(row.get("sku") or "").strip(),
            "name": str(row.get("name") or "").strip(),
            "category": str(row.get("category") or "").strip(),
            "stock": row.get("stock"),
        }
        try:
            sku = normalize_sku(result["sku"])
            name = str(row.get("name") or "").strip()
            description = str(row.get("description") or "").strip()
            category_name = canonical_category_name(str(row.get("category") or "").strip())
            operation = _parse_import_operation(row.get("operation"))
            active = _parse_import_active(row.get("active"))
            price = _parse_import_decimal(row.get("price"), field_name="price")
            stock = _parse_import_stock(row.get("stock"))

            if not name:
                raise HTTPException(status_code=400, detail="name is required")
            if not description:
                raise HTTPException(status_code=400, detail="description is required")
            if not category_name:
                raise HTTPException(status_code=400, detail="category is required")

            product = existing_products.get(sku)
            requested_action = operation or ("archive" if not active else "upsert")
            category_exists = category_name.casefold() in categories_by_name

            if requested_action == "archive":
                if not product:
                    raise HTTPException(status_code=400, detail="cannot archive unknown SKU")
                action = "skip" if product.archived_at else "archive"
                reason = "already archived" if product.archived_at else None
            elif product:
                action = "update"
                reason = "will restore archived product" if product.archived_at else None
            else:
                action = "create"
                reason = None

            result.update(
                {
                    "sku": sku,
                    "name": name,
                    "description": description,
                    "price": money_json(price),
                    "category": category_name,
                    "stock": 0 if stock is None and action == "create" else stock,
                    "active": active,
                    "operation": requested_action,
                    "action": action,
                    "category_action": "existing" if category_exists else "create",
                    "message": reason,
                }
            )
            summary[action] += 1
        except HTTPException as exc:
            result["action"] = "error"
            result["message"] = exc.detail
            summary["error"] += 1
        preview_rows.append(result)

    return {"summary": summary, "rows": preview_rows}


def _apply_import(preview: dict, db: Session, *, filename: str, created_by: str) -> dict:
    if preview["summary"]["error"]:
        raise HTTPException(status_code=400, detail="Fix import errors before apply")

    categories_by_name = {
        category.name.casefold(): category
        for category in db.query(Category).all()
    }
    inventory_updates: list[dict] = []
    applied = {"created": 0, "updated": 0, "archived": 0, "restored": 0, "stock_updates": 0}

    for row in preview["rows"]:
        if row["action"] == "skip":
            continue

        category = categories_by_name.get(row["category"].casefold())
        if not category:
            category = Category(name=row["category"], description=f"Importată din Excel pe {datetime.utcnow().date().isoformat()}")
            db.add(category)
            db.flush()
            categories_by_name[category.name.casefold()] = category

        product = db.query(Product).filter(Product.sku == row["sku"]).first()
        if row["action"] == "archive":
            assert product is not None
            _set_product_archived_at(product, datetime.utcnow())
            db.add(product)
            inventory_updates.append({"product_id": _product_id(product), "stock": 0})
            applied["archived"] += 1
            applied["stock_updates"] += 1
            continue

        if product:
            restored = product.archived_at is not None
            product.name = row["name"]
            product.description = row["description"]
            product.price = as_money(row["price"])
            product.category_id = category.id
            _set_product_archived_at(product, None)
            db.add(product)
            applied["updated"] += 1
            if restored:
                applied["restored"] += 1
        else:
            product = Product(
                sku=row["sku"],
                name=row["name"],
                description=row["description"],
                price=as_money(row["price"]),
                category_id=category.id,
                archived_at=None,
            )
            db.add(product)
            db.flush()
            applied["created"] += 1

        if row["stock"] is not None:
            inventory_updates.append({"product_id": _product_id(product), "stock": int(row["stock"])})
            applied["stock_updates"] += 1

    db.commit()
    if inventory_updates:
        _sync_inventory_bulk(inventory_updates)
    redis_client.delete(PRODUCT_CACHE_KEY)
    result = {
        "message": "Import completed",
        "summary": applied,
        "rows": preview["rows"],
    }
    db.add(
        ProductImportJob(
            filename=filename[:255],
            summary_json=json.dumps(result["summary"]),
            created_by=created_by[:255],
        )
    )
    db.commit()
    return result


@app.get("/categories")
def list_categories(db: Session = Depends(get_db)):
    return {"items": [serialize_category(category) for category in db.query(Category).order_by(Category.name).all()]}


@app.get("/products")
def list_products(
    include_archived: bool = Query(default=False),
    db: Session = Depends(get_db),
):
    cached = redis_client.get(PRODUCT_CACHE_KEY)
    if cached and not include_archived:
        return {"source": "cache", "items": json.loads(cached)}

    categories_by_id = {cast(int, category.id): category for category in db.query(Category).all()}
    query = db.query(Product)
    if not include_archived:
        query = query.filter(Product.archived_at.is_(None))
    products = [serialize_product(product, categories_by_id) for product in query.all()]
    if not include_archived:
        redis_client.setex(PRODUCT_CACHE_KEY, 120, json.dumps(products))
    return {"source": "database", "items": products}


@app.get("/products/export")
def export_products_excel(
    include_archived: bool = Query(default=True),
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    from openpyxl import Workbook

    categories_by_id = {cast(int, category.id): category for category in db.query(Category).all()}
    query = db.query(Product)
    if not include_archived:
        query = query.filter(Product.archived_at.is_(None))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Produse"
    sheet.append(
        [
            "sku",
            "name",
            "description",
            "price",
            "category",
            "active",
            "archived_at",
        ]
    )
    for product in query.order_by(Product.id).all():
        category = categories_by_id.get(_product_category_id(product))
        archived_at = _product_archived_at(product)
        sheet.append(
            [
                _product_sku(product),
                _product_name(product),
                _product_description(product),
                money_json(product.price),
                category.name if category else "",
                "true" if archived_at is None else "false",
                archived_at.isoformat() if archived_at else "",
            ]
        )

    payload = BytesIO()
    workbook.save(payload)
    payload.seek(0)
    return StreamingResponse(
        payload,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="magazon-products.xlsx"'},
    )


@app.get("/products/{product_id}")
def get_product(product_id: int, db: Session = Depends(get_db)):
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product or product.archived_at is not None:
        raise HTTPException(status_code=404, detail="Product not found")
    categories_by_id = {cast(int, category.id): category for category in db.query(Category).all()}
    return serialize_product(product, categories_by_id)


@app.post("/products")
def create_product(
    payload: ProductRequest,
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    if payload.category_id is not None:
        category = db.query(Category).filter(Category.id == payload.category_id).first()
        if not category:
            raise HTTPException(status_code=404, detail="Category not found")
    values = payload.model_dump()
    values["price"] = as_money(values["price"])
    requested_sku = values.pop("sku", None)
    if requested_sku:
        requested_sku = normalize_sku(requested_sku)
        if db.query(Product).filter(Product.sku == requested_sku).first():
            raise HTTPException(status_code=409, detail="SKU already exists")
    values["sku"] = requested_sku or f"TEMP-{uuid4()}"
    product = Product(**values)
    db.add(product)
    db.flush()
    if not requested_sku:
        _set_product_sku(product, generate_sku(_product_name(product), _product_id(product)))
    db.commit()
    db.refresh(product)
    redis_client.delete(PRODUCT_CACHE_KEY)
    categories_by_id = {cast(int, category.id): category for category in db.query(Category).all()}
    return serialize_product(product, categories_by_id)


@app.put("/products/{product_id}")
def update_product(
    product_id: int,
    payload: ProductRequest,
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    if payload.category_id is not None:
        category = db.query(Category).filter(Category.id == payload.category_id).first()
        if not category:
            raise HTTPException(status_code=404, detail="Category not found")

    values = payload.model_dump()
    values["price"] = as_money(values["price"])
    requested_sku = values.pop("sku", None)
    if requested_sku:
        requested_sku = normalize_sku(requested_sku)
        existing = db.query(Product).filter(
            Product.sku == requested_sku,
            Product.id != product_id,
        ).first()
        if existing:
            raise HTTPException(status_code=409, detail="SKU already exists")
        values["sku"] = requested_sku
    for field, value in values.items():
        setattr(product, field, value)
    _set_product_archived_at(product, None)
    db.add(product)
    db.commit()
    db.refresh(product)
    redis_client.delete(PRODUCT_CACHE_KEY)
    categories_by_id = {cast(int, category.id): category for category in db.query(Category).all()}
    return serialize_product(product, categories_by_id)


@app.delete("/products/{product_id}")
def delete_product(
    product_id: int,
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    db.delete(product)
    db.commit()
    redis_client.delete(PRODUCT_CACHE_KEY)
    return {"message": "Product deleted", "id": product_id}


@app.post("/categories")
def create_category(
    payload: CategoryRequest,
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    name = canonical_category_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="Category name is required")
    existing = db.query(Category).filter(func.lower(Category.name) == name.lower()).first()
    if existing:
        raise HTTPException(status_code=409, detail="Category already exists")
    category = Category(name=name, description=payload.description.strip())
    db.add(category)
    db.commit()
    db.refresh(category)
    redis_client.delete(PRODUCT_CACHE_KEY)
    return serialize_category(category)


@app.get("/products/import/template")
def download_import_template(_admin=Depends(require_admin)):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Produse"
    sheet.append(["sku", "name", "description", "price", "category", "stock", "active", "operation"])
    sheet.append(["DOCK-USBC-001", "USB-C Dock", "Dock cu HDMI si ethernet", 149.99, "Accesorii", 25, "true", "upsert"])
    sheet.append(["OLD-MOUSE-001", "Mouse vechi", "Produs retras", 49.00, "Periferice", 0, "false", "archive"])
    payload = BytesIO()
    workbook.save(payload)
    payload.seek(0)
    return StreamingResponse(
        payload,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="magazon-import-template.xlsx"'},
    )


@app.get("/products/import/jobs")
def list_product_import_jobs(
    limit: int = Query(default=20, ge=1, le=100),
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    jobs = (
        db.query(ProductImportJob)
        .order_by(ProductImportJob.created_at.desc(), ProductImportJob.id.desc())
        .limit(limit)
        .all()
    )
    return {"items": [serialize_import_job(job) for job in jobs], "total": len(jobs)}


@app.post("/products/import/preview")
async def preview_product_import(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin_claims: dict = Depends(require_admin),
):
    client_ip = request.client.host if request.client else "unknown"
    enforce_rate_limit(f"product:import-preview:{client_ip}", limit=20, window_seconds=900)
    file_bytes = await file.read()
    _enforce_import_upload(file, file_bytes)
    preview = _build_import_preview(_load_import_rows(file_bytes), db)
    return {"message": "Preview generated", **preview}


@app.post("/products/import/apply")
async def apply_product_import(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin_claims: dict = Depends(require_admin),
):
    client_ip = request.client.host if request.client else "unknown"
    enforce_rate_limit(f"product:import-apply:{client_ip}", limit=10, window_seconds=900)
    file_bytes = await file.read()
    _enforce_import_upload(file, file_bytes)
    preview = _build_import_preview(_load_import_rows(file_bytes), db)
    created_by = admin_claims.get("email") or admin_claims.get("sub") or "admin"
    return _apply_import(preview, db, filename=file.filename or "upload.xlsx", created_by=created_by)
